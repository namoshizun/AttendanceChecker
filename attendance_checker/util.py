import time, sys, json
import numpy as np
import pandas as pd
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows

#############
# MISC UTIL #
#############
def read_json(path):
    with open('config.json', 'r') as file:
        return json.load(file)


def load_workbook(path):
    def loader(wb, sheet_idx=0):
        try:
            sheet_names = wb.get_sheet_names()
            curr = sheet_names[sheet_idx]
        except IndexError:
            return {}
        # construct dataframe from worksheet (only gets member's number, actual name and nickname)
        ws = wb.get_sheet_by_name(curr)
        df = pd.DataFrame(ws.values).iloc[:, :3]
        # clean data
        df.dropna(inplace=True)
        df.columns = ['编号', '姓名', 'YY昵称', '截屏1', '截屏2', '结果']
        df['YY昵称'] = df['YY昵称'].apply(str.strip)
        # use YY names as index
        df.set_index('YY昵称', inplace=True)
        # placeholders for makring attendance
        polyfill = [0] * len(df)
        df['截屏1'], df['截屏2'], df['结果'] = polyfill, polyfill, polyfill
        # recursively build from sheets
        data = {curr: df}
        data.update(loader(wb, sheet_idx + 1))
        return data
    
    wb = pyxl.load_workbook(path)
    return loader(wb)


def save_workbook(member_sheet):
    wb = pyxl.Workbook(write_only=True)
    for sheet, df in member_sheet.data.items():
        ws = wb.create_sheet(sheet)
        next(map(ws.append, list(dataframe_to_rows(df))[1:]))
        # for row in list(dataframe_to_rows(df))[1:]:
        #     ws.append([row[1], row[2], row[0], row[3], row[4], row[5]])  # revert back to original order
    wb.save(member_sheet.source)


######################
# TEXT PREPROCESSING #
######################
DEFAULT_ENC = sys.getdefaultencoding()
def select_encoding(ioFn):
    encodings = ['utf-8', 'utf-16', 'gb18030', 'gb2312', 'gbk']

    def detect_encoding(path, idx=0):
        try:
            with open(path, encoding=encodings[idx]) as source:
                source.read()
            return True, encodings[idx]
        except IndexError:
            return False, None
        except UnicodeDecodeError:
            return detect_encoding(path, idx+1)


    def checker(self, path):
        succ, encoding = detect_encoding(path)
        if succ:
            return ioFn(self, path, encoding)
        else:
            raise UnicodeDecodeError('cannot decode file ' + path)
    return checker

#######################
# IMAGE PREPROCESSING #
#######################
def rgb2gray(rgb):
    return np.dot(rgb[...,:3], [0.3, 0.59, 0.11])


def gray2wb(gray, normalise=True):
    bools = gray >= 25.
    gray[bools] = 255 if not normalise else 1.0
    gray[np.invert(bools)] = 0.0
    
    return gray


def draw(*args):
    import matplotlib.pyplot as plt
    for i, img in enumerate(args):
        plt.subplot(2, 1, i + 1)
        plt.imshow(img)
    plt.show()


def first_occurrence(arr, val):
    for i, v in enumerate(arr):
        if v == val:
            return i

    return -1
